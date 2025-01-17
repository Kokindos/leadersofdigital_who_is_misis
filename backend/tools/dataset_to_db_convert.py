import os
import time

import geoalchemy2.shape
import openpyxl
import shapefile
import shapely.geometry
import sqlalchemy.exc
from alive_progress import alive_bar
from shapely import geometry, wkt, wkb

from backend import db, app
from backend.data.models import *
from backend.data.models.start_ground import StartGround
from backend.services.polygons import split_multipoint_by_parts
from backend.tools import coord_convert

converter = coord_convert.CoordConverter()

DATASET_PATHS = {
    "lands": "ЗУ/Земельные_участки.shp",
    "capital": "ОКС/ОКС.shp",
    "organizations": "Организации СВАО_САО/Организации_СВАО_САО.shp",
    "protected_zones": "Санитарно-защитные зоны/СЗЗ.shp",
    "start_grounds": "Стартовые площадки/Стартовые_площадки_реновации.shp",
    "cultural_inheritance": "Территории объектов культурного наследия/Территории_объектов_культурного_наследия.shp",
    "bad_records": "Аварийные_Самовольные_Несоответствие_ВРИ_СВАО_САО.XLSX",
    "buildings": "Здания СВАО_САО жилое_нежилое.xlsx"
}
TRUNCATE_TABLE = "TRUNCATE TABLE {} CASCADE"
YES = "Да"
HABITABLE_STR = "жилое"


def main(path_to_dataset, progress_output_stream=None):
    i = 0

    def progress_log_wrapper(f, *args, _ostream=progress_output_stream, _caption, **kwargs):
        nonlocal i
        i += 1

        try:
            start = time.time()
            f(*args, **kwargs)
            if _ostream is not None:
                _ostream.write(f"({i / 8 * 100}%)\t -- \"{_caption}\" loaded in {time.time() - start : .3f}s\n")
        except Exception as e:
            if _ostream is not None:
                _ostream.write(f"\"{_caption}\" cannot be loaded, all changes cancelled\n")
            raise e
        return None

    progress_log_wrapper(load_lands, path_to_dataset, _caption="ЗУ")
    progress_log_wrapper(load_capital, path_to_dataset, _caption="ОКС")
    progress_log_wrapper(load_organizations, path_to_dataset, _caption="Организации СВАО_САО")
    progress_log_wrapper(load_protected_zones, path_to_dataset, _caption="Санитарно-защитные зоны")
    progress_log_wrapper(load_start_grounds, path_to_dataset, _caption="Стартовые площадки")
    progress_log_wrapper(load_cultural_inheritance, path_to_dataset,
                         _caption="Территории объектов культурного наследия")
    progress_log_wrapper(load_bad_buildings, path_to_dataset,
                         _caption="Аварийные_Самовольные_Несоответствие_ВРИ_СВАО_САО")
    progress_log_wrapper(load_buildings, path_to_dataset, _caption="Здания СВАО_САО жилое_нежилое")

    db.session.commit()

    for file in os.listdir(app.config["PREPROCESSED_DATA_PATH"]):
        os.remove(os.path.join(app.config["PREPROCESSED_DATA_PATH"], file))


def convert_bbox(x1, y1, x2, y2):
    x1, y1 = converter.transform(x1, y1, coord_convert.CRS_MSK)
    x2, y2 = converter.transform(x2, y2, coord_convert.CRS_MSK)
    return y1, x1, y2, x2


def convert_multipoint(points):
    return list(map(lambda p: converter.transform(p[0], p[1], coord_convert.CRS_MSK)[::-1], points))


def load_lands(path_to_dataset):  # ЗУ
    db.session.execute(TRUNCATE_TABLE.format(Land.__tablename__))

    reader = shapefile.Reader(os.path.join(path_to_dataset, DATASET_PATHS["lands"]), encoding="cp1251")
    with alive_bar(len(reader.shapeRecords())) as bar:
        for shaperec in reader.shapeRecords():
            splited_multipoints = split_multipoint_by_parts(
                list(map(lambda x: shapely.geometry.Point(*x), convert_multipoint(shaperec.shape.points))),
                shaperec.shape.parts)
            obj = Land(
                oid=shaperec.shape.oid,
                parts=list(shaperec.shape.parts),
                points=wkt.dumps(geometry.MultiPolygon(list(map(lambda x: geometry.Polygon(x), splited_multipoints)))),
                bbox=wkt.dumps(geometry.box(*convert_bbox(*shaperec.shape.bbox))),

                cadnum=shaperec.record.DESCR,
                address=shaperec.record.ADDRESS,
                has_effect=shaperec.record.HAS_EFFECT,
                property_t=shaperec.record.PROPERTY_T,
                shape_area=shaperec.record.SHAPE_AREA
            )
            db.session.add(obj)
            bar()
    reader.close()


def load_capital(path_to_dataset):  # ОКС

    db.session.execute(TRUNCATE_TABLE.format(CapitalConstructionWorks.__tablename__))
    reader = shapefile.Reader(os.path.join(path_to_dataset, DATASET_PATHS["capital"]), encoding="cp1251")
    with alive_bar(len(reader.shapeRecords())) as bar:
        for shaperec in reader.shapeRecords():
            splited_multipoints = split_multipoint_by_parts(
                list(map(lambda x: shapely.geometry.Point(*x), convert_multipoint(shaperec.shape.points))),
                shaperec.shape.parts)

            obj = CapitalConstructionWorks(
                oid=shaperec.shape.oid,
                parts=list(shaperec.shape.parts),
                points=wkt.dumps(geometry.MultiPolygon(list(map(lambda x: geometry.Polygon(x), splited_multipoints)))),

                bbox=wkt.dumps(geometry.box(*convert_bbox(*shaperec.shape.bbox))),

                cadnum=shaperec.record.cadnum,
                address=shaperec.record.address,
                area=shaperec.record.Area
            )
            db.session.add(obj)
            bar()
    reader.close()


def load_organizations(path_to_dataset):  # Организации СВАО_САО
    db.session.execute(TRUNCATE_TABLE.format(Organization.__tablename__))

    reader = shapefile.Reader(os.path.join(path_to_dataset, DATASET_PATHS["organizations"]), encoding="cp1251")
    with alive_bar(len(reader.shapeRecords())) as bar:
        for shaperec in reader.shapeRecords():
            obj = Organization(
                oid=shaperec.shape.oid,
                point=wkt.dumps(
                    geometry.Point(converter.transform(*shaperec.shape.points[0], coord_convert.CRS_MSK)[::-1])),

                name=shaperec.record.name,
                kol_mest=shaperec.record.kol_mest
            )
            db.session.add(obj)
            bar()
    reader.close()


def load_protected_zones(path_to_dataset):  # Санитарно-защитные зоны
    db.session.execute(TRUNCATE_TABLE.format(SanitaryProtectedZone.__tablename__))

    reader = shapefile.Reader(os.path.join(path_to_dataset, DATASET_PATHS["protected_zones"]), encoding="cp1251")
    with alive_bar(len(reader.shapeRecords())) as bar:
        for shaperec in reader.shapeRecords():
            splited_multipoints = split_multipoint_by_parts(
                list(map(lambda x: shapely.geometry.Point(*x), convert_multipoint(shaperec.shape.points))),
                shaperec.shape.parts)

            obj = SanitaryProtectedZone(
                oid=shaperec.shape.oid,
                parts=list(shaperec.shape.parts),
                points=wkt.dumps(geometry.MultiPolygon(list(map(lambda x: geometry.Polygon(x), splited_multipoints)))),
                bbox=wkt.dumps(geometry.box(*convert_bbox(*shaperec.shape.bbox))),

                zone_type=shaperec.record.VID_ZOUIT
            )
            db.session.add(obj)
            bar()
    reader.close()


def load_start_grounds(path_to_dataset):  # Стартовые площадки
    db.session.execute(TRUNCATE_TABLE.format(StartGround.__tablename__))

    reader = shapefile.Reader(os.path.join(path_to_dataset, DATASET_PATHS["start_grounds"]), encoding="cp1251")
    with alive_bar(len(reader.shapeRecords())) as bar:
        for shaperec in reader.shapeRecords():
            splited_multipoints = split_multipoint_by_parts(
                list(map(lambda x: shapely.geometry.Point(*x), convert_multipoint(shaperec.shape.points))),
                shaperec.shape.parts)

            obj = StartGround(
                oid=shaperec.shape.oid,
                parts=list(shaperec.shape.parts),
                points=wkt.dumps(geometry.MultiPolygon(list(map(lambda x: geometry.Polygon(x), splited_multipoints)))),
                bbox=wkt.dumps(geometry.box(*convert_bbox(*shaperec.shape.bbox))),

                address=shaperec.record.address,
                district=shaperec.record.rayon
            )
            db.session.add(obj)
            bar()
    reader.close()


def load_cultural_inheritance(path_to_dataset):  # Территории объектов культурного наследия
    db.session.execute(TRUNCATE_TABLE.format(CulturalHeritage.__tablename__))

    reader = shapefile.Reader(os.path.join(path_to_dataset, DATASET_PATHS["cultural_inheritance"]), encoding="cp1251")
    with alive_bar(len(reader.shapeRecords())) as bar:
        for shaperec in reader.shapeRecords():
            splited_multipoints = split_multipoint_by_parts(
                list(map(lambda x: shapely.geometry.Point(*x), convert_multipoint(shaperec.shape.points))),
                shaperec.shape.parts)

            obj = CulturalHeritage(
                oid=shaperec.shape.oid,
                parts=list(shaperec.shape.parts),
                points=wkt.dumps(geometry.MultiPolygon(list(map(lambda x: geometry.Polygon(x), splited_multipoints)))),
                bbox=wkt.dumps(geometry.box(*convert_bbox(*shaperec.shape.bbox))),

                name=shaperec.record.name,
                number=shaperec.record.number,
                object_id=shaperec.record.objectid

            )
            db.session.add(obj)
            bar()
    reader.close()


def load_bad_buildings(path_to_dataset):
    db.session.execute(TRUNCATE_TABLE.format(BadBuilding.__tablename__))

    wb = openpyxl.load_workbook(filename=os.path.join(path_to_dataset, DATASET_PATHS["bad_records"]))
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        district, cadnum, unauthorized, mismatch, hazardous = row
        obj = BadBuilding(
            cadnum=cadnum,
            district=district,
            unauthorized=unauthorized == YES,
            mismatch=mismatch == YES,
            hazardous=hazardous == YES,
        )
        db.session.merge(obj)
    wb.close()


def load_buildings(path_to_dataset):
    db.session.execute(TRUNCATE_TABLE.format(Building.__tablename__))

    wb = openpyxl.load_workbook(filename=os.path.join(path_to_dataset, DATASET_PATHS["buildings"]))
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        cadnum, address, area, habitable, year_built, wall_material, hazardous, typical = row
        if year_built == '' or year_built == 0:
            year_built = None

        obj = Building(
            cadnum=cadnum,
            address=address,
            area=area,
            habitable=habitable == HABITABLE_STR,
            year_built=year_built,
            wall_material=wall_material,
            hazardous=hazardous == YES,
            typical=typical == YES
        )
        db.session.merge(obj)
    wb.close()


def load_extended_lands():
    db.session.execute(TRUNCATE_TABLE.format(ExtendedLand.__tablename__))
    db.session.execute(TRUNCATE_TABLE.format(ExtendedCapitalConstructionWorks.__tablename__))
    db.session.execute(TRUNCATE_TABLE.format(ExtendedOrganization.__tablename__))
    db.session.commit()


    all_lands = db.session.execute("select *, st_asewkt(points) as wktpoints from lands").all()
    all_sgs = db.session.execute("select *, st_asewkt(points) as wktpoints from start_grounds").all()

    with alive_bar(len(all_lands) + len(all_sgs)) as bar:
        for land in all_lands:
            obj = ExtendedLand(
                land_oid=land.oid,
                start_ground_oid=None,

                is_sanitary_protected_zone=check_if_land_is_sanitary_protected_zone(land.wktpoints),
                is_cultural_heritage=check_if_land_is_cultural_heritage(land.wktpoints),

                **get_bad_building_data(land),

            )
            db.session.add(obj)
            db.session.commit()
            db.session.refresh(obj)
            load_extended_capital_construction_works(obj, land.wktpoints)
            bar()

        #         # capital_construction_works = db.session.query(CapitalConstructionWorks).select(.points=)

        for start_ground in all_sgs:
            obj = ExtendedLand(
                land_oid=None,
                start_ground_oid=start_ground.oid,

                is_sanitary_protected_zone=check_if_land_is_sanitary_protected_zone(start_ground.wktpoints),
                is_cultural_heritage=check_if_land_is_cultural_heritage(start_ground.wktpoints),

            )
            db.session.add(obj)
            db.session.commit()
            db.session.refresh(obj)
            load_extended_capital_construction_works(obj, start_ground.wktpoints)
            bar()


def get_bad_building_data(land: Land) -> dict:
    bad_building = db.session.query(BadBuilding).get(land.cadnum)
    bad_building_data = {
        'is_unauthorized': None,
        'is_hazardous': None,
        'is_mismatch': None,

    }
    if bad_building is None:
        return bad_building_data
    else:
        bad_building_data = {
            'is_unauthorized': bad_building.unauthorized,
            'is_hazardous': bad_building.hazardous,
            'is_mismatch': bad_building.mismatch
        }
    return bad_building_data


def load_extended_capital_construction_works(extended_land, land_wktpoints: str):

    for oks in db.session.execute(f"select *, st_asewkt(points) as wktpoints from capital_construction_works "
                                  f"where st_intersects(points, '{land_wktpoints}');").all():
        obj = ExtendedCapitalConstructionWorks(
            oid=oks.oid,
            extended_land_id=extended_land.id,

            **get_building_data(oks)
        )
        db.session.merge(obj)
        db.session.commit()
        load_extended_organizations(obj, oks.wktpoints)


def get_building_data(oks: CapitalConstructionWorks) -> dict:
    building = Building.query.get(oks.cadnum)
    building_data = {
        'habitable': None,
        'area': None,
        'year_built': None,
        'wall_material': None,
        'hazardous': None,
        'typical': None,
    }

    if building is None:
        return building_data

    building_data = {
        'habitable': building.habitable,
        'area': building.area,
        'year_built': building.year_built,
        'wall_material': building.wall_material,
        'hazardous': building.hazardous,
        'typical': building.typical,
    }
    return building_data


def load_extended_organizations(extended_oks, oks_wktpoints):
    for org in db.session.execute(f"select * from organizations "
                                  f"where st_contains('{oks_wktpoints}', point);").all():
        obj = ExtendedOrganization(
            oid=org.oid,
            extended_capital_construction_works_oid=extended_oks.oid,
        )
        db.session.merge(obj)
        db.session.commit()


def check_if_land_is_sanitary_protected_zone(land_wktpoints: str) -> bool:
    zones_intersected = db.session.execute(f"select oid from sanitary_protected_zones where "
                                           f"st_intersects(points, '{land_wktpoints}');").all()
    return len(zones_intersected) > 0


def check_if_land_is_cultural_heritage(land_wktpoints: str) -> bool:
    zones_intersected = db.session.execute(f"select oid from cultural_heritage where "
                                           f"st_intersects(points, '{land_wktpoints}');").all()
    return len(zones_intersected) > 0
